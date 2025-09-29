#!/usr/bin/env python3
"""F1 Analytics: Visualizations and Export to Excel."""

import pandas as pd
import plotly.express as px
from connection import F1DatabaseConnector
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
import os

# Directories
CHARTS_DIR = os.path.join(os.path.dirname(__file__), "charts")
EXPORTS_DIR = os.path.join(os.path.dirname(__file__), "exports")

def get_top_drivers_by_season(connector):
    """Get top 10 drivers by points per season, excluding those with 0 points."""
    query = """
    SELECT ra.year, d.forename || ' ' || d.surname AS driver, SUM(r.points) AS total_points
    FROM results r
    JOIN drivers d ON d.driverid = r.driverid
    JOIN races ra ON ra.raceid = r.raceid
    WHERE r.points > 0
    GROUP BY ra.year, d.driverid, driver
    ORDER BY ra.year, total_points DESC;
    """
    df = pd.read_sql_query(query, connector.connection)
    # For each year, keep only top 10
    df_top = df.groupby('year').head(10).reset_index(drop=True)
    return df_top

def create_plotly_animation(df):
    """Create interactive Plotly chart with time slider for top drivers by season."""
    fig = px.bar(df, x='driver', y='total_points', animation_frame='year',
                 title='Top 10 F1 Drivers by Points per Season',
                 labels={'total_points': 'Total Points', 'driver': 'Driver'},
                 color='driver')
    fig.update_layout(xaxis_tickangle=-45)
    # Save to HTML for viewing
    os.makedirs(CHARTS_DIR, exist_ok=True)
    fig.write_html(os.path.join(CHARTS_DIR, 'top_drivers_by_season.html'))
    print("Interactive chart saved to charts/top_drivers_by_season.html")
    fig.show()

def export_to_excel(dataframes_dict, filename):
    """Export dataframes to Excel with formatting."""
    os.makedirs(EXPORTS_DIR, exist_ok=True)
    filepath = os.path.join(EXPORTS_DIR, filename)
    
    total_sheets = 0
    total_rows = 0
    
    # Write data to Excel
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            total_sheets += 1
            total_rows += len(df)
    
    # Apply formatting
    workbook = load_workbook(filepath)
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Frozen rows/columns for headers
        worksheet.freeze_panes = "B2"
        
        # Filters on all columns
        worksheet.auto_filter.ref = worksheet.dimensions
        
        # Conditional formatting (gradient fill for numeric columns)
        for col in worksheet.iter_cols(min_row=2):
            col_letter = col[0].column_letter
            # Check if column contains numeric data
            if any(cell.data_type in ('n', 'f') for cell in col if cell.value is not None):
                # Gradient fill for numeric columns
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",  # Red for min
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",  # Yellow for mid  
                    end_type="max", end_color="FF00AA00"  # Green for max
                )
                range_ref = f"{col_letter}2:{col_letter}{worksheet.max_row}"
                worksheet.conditional_formatting.add(range_ref, rule)
    
    workbook.save(filepath)
    print(f"Created file {filename}, {total_sheets} sheets, {total_rows} rows")
    return filepath

def main():
    print("F1 Analytics: Visualizations and Export")
    connector = F1DatabaseConnector()
    if not connector.connect():
        print("Failed to connect to database.")
        return
    try:
        df = get_top_drivers_by_season(connector)
        print(f"Loaded {len(df)} rows of data.")
        print("Created interactive bar chart: Top 10 F1 Drivers by Points per Season (with time slider)")
        print("The chart shows the evolution of top drivers' points across seasons, excluding drivers with 0 points.")
        
        # Create Plotly visualization
        create_plotly_animation(df)
        
        # Export to Excel with formatting
        print("\nExporting data to Excel...")
        dataframes_dict = {
            'Top_Drivers_by_Season': df
        }
        export_to_excel(dataframes_dict, 'f1_analytics_report.xlsx')
        
        print("\n✅ Visualization and export completed successfully!")
        
    finally:
        connector.disconnect()

if __name__ == "__main__":
    main()